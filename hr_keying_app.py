import streamlit as st
import os
import json
import io
import base64
import requests
import pytesseract
from openpyxl import load_workbook
import pandas as pd
from PIL import Image

# โลโก้ AREE Workforce Tech (ฝังเป็น base64 เพื่อให้แสดงผลได้แน่นอนไม่ต้องพึ่งไฟล์ภายนอกตอน deploy)
LOGO_BASE64 = "iVBORw0KGgoAAAANSUhEUgAAAIAAAACACAYAAADDPmHLAAAQU0lEQVR42u2dfXBc1XnGn/c992MlWbYiG1vy11qOAraEIQyF0CQgm+DBIZRJw6zSZEjTodNkmDbQaYBQXCOJfNCkoUnI5I90OuSjUybjpQNpS8vQplhJE0MKBLDBwTE0EjY22GDLtSRr95737R/3riWMZGuvbWl3fX6aO7ra3bl7dc/7POc9H/dcwOFwOBwOh8PhcDgcDofD4XA4HA6Hw+FwOBwOh8PhcDgcDofD4XA4HA6Hw+FwOBwOh8NRNdDZ+o8rQL3oMp1YqC8grwDQCdA56KIt6Jc+QFx41CibAXPyz+TM2XAtvLOt8HsA7gbs11vPXTDPx0d9pSt86LJAiOuJ9jWo+cUcyTz0O3vzgwBIY5tUVwXUwP/aA1AfIF9f1n5TPeiuRjYtPgCjhECBDBHqQIDgcKj8jUv3PN9HINRyEJw1AbAZMN2A/cqy9vua2XzOqoJVoxCAB6ZAgRCAD2gI9haxh0JED1/cell3/umDkkNeajEIzNlU+L3LVt4yz3g9o9YWGSAPZAyIDYg9IPlNTCAdE1tc4JnOgaHXmtcf+a9HOpEzebyozgGqsM7vBbR32XmthqKdISjDCg5BFIAQAPARVwET//YBDRW2kXwPUcOla/ZtfWozctyNvK2l68O1HgCdyBEBGlF0e8imwaoqTSPwCSAByCNFRGO9cDlAdaq/D9Dbl7QvCYz+OgTVGYACRaL+qR0gAOArwYdKI+pZi3N+t+ON/ic2I2dqyQW4ttUPAqDKcnvI3CBQofKDXn0CilzsrcVrxLWs/m5Abm5duZwIfzwmIpou6TVHJJI6Nldva/3gB7qRt7XUScS1rn42uCNkrk+p/lJNqR4IVqXHOUAVqf+WbHYFGH80JiIESq1agmf+TyOpZ2/9C61XXF5LLsC1rP6C8h0hc53EAzuUrvANSD0AqgaESKUXAHLoUBcAFaz+m5a/ZyUDn47Vn7bDS2E0AMU/5ohGUk/eldsWda0l9EktuEDNBcCLJfVT8U7fcEZVU6lfATAMPPUndgAoA7CoHRfgWlN/HpAbli5tZ+CGgoiAKLX6ffVBcTwl1UHsAnPY69q+aN2VhD7RKncBrkX1E3sbfTahavq638AgUA9aGgs8lhNACYQocQFUuQtwran/421t5zLhkwWxgpR1v0KROab+d6SFZlgj28je5c+3Xrm+2l2Aa079Vjb6zIECMkUJnrTu98DIHFP/CT4r1e8CXEvq/1g2u4oInyiInJL6G9Q/2YVJXMB///aWD22oZhfgWlK/qm7ymH0FUvX6KQAfjHo105oRqlAItEcBqlYX4FpR/3XLl3cwUXfxFNr9CsVc9cDTiB0CmRGN7BzyLnux5aoPV6sLcK2oX4BNHrOHFL1+pYZeCEaDepAyhv4VQCTSW60uUNUBkANMHpAPr1y6hhi5Ylz3e+nUDzSpAaOcmR9khiWyczm4ZPui9ddWowvUQg6g1vImw2wQ9/qVjYVqBiRzk7q/HPsgAgSqqtrTgx6uNhfgalf/uhVLLiTGxyIrAiIvjfI9ELXCY0rygPIgM6yRNLJ/8fULt15H6JPH0eW5AJgh9avwXYbYaKo5eyp1xFqEPlkn9Eg9Mwgoe7oXJS6ihB5FD6/FWnEBMAPqvyKbvYgMfTQSEaLyM38F1CeiIuvdLSa4JdJj5Vm2C4xoJPPIf+/2lq2/X00uUM0OoEXYHkPECk1T79oMMQ9JtK19YPFji159+uURsf/cyIY1pQtEEBXFXZuRM9XiAlyt6n/f8uUXM9N1UcrZPgqFISIRvnsd+iMFKCD0jmnaqWOJC3BwwapFh6+vFheoVgdQS7aXiShd3Q+bIcPDNnq2ac9LD8WdSTnufO35X42q/LiRDQMapQgCKqoolDY9ji6vGlyAq1D99sK2pZcS80dsil4/TTYmEEB3dwO2EzlK3iOf0HdUxRKo7GtDAI9qJPOMf37zwkyuGlygKh1AVXqYiJBG/RrX/SNWntm3+zc/jqeQ5W038jaPHK/Z8+xzYyIPxS6AVC5QUFGANim6vC3oFxcAp1H9q7JLLmPma2zqET8FE0hI+voAKal/gkMQGb57VK2lFNeHQDyqkTRxuHrbwoY/6AMq2gWqzwFYeym+W1/LzdUUsCEbM2LlqT9/9eV/Kam/9H7JBS7c/fS2o6IPNrKXOhcoqFVSbFR8xk9cgFwApMfkAbtyZesHmPhqKyKgdCN+cWVPvQRo5ySFkkOHKkCBx18cVRvFuYCmcAErczlYtf2cPZ9IXMC4ADhFSNCbFFnZdb9CbUhsRsQ+edue3/xbaamYdxZen+SR485X/+eFMZHNsQukyQVAY7CqRH+1HblgC/pFK9AFqiEADAC7rK31cjBfJaLp1U8AM3oxhfqPdwED86VRscWkRaDlfRfxUbUyj/33yIKRG/oA2VKBLlANAaAAQIq+OPFP1+4PyZhRa7feOfDyo1Op/3gXWLPvyR0F6I+SfgGbygXUKpju3Nm+IaxEF+AqUL9ks0vWGuZ1cgpz/eL1vrgXOHbr2ElaHXlVgED2y6MihSQXKM8FELtAEwfvjg6Hn6pEF+BqUD9Yeill3Q+oDZjNUZH/7nt112MnU/+EZFHyyPGFrz310hjkgUb2GIpULnBUrVrVO/8325WpNBfgSlf/8pWLP8TEXVY0lfoVyTpvwLTVf7wLeBG+PCK2wJQiFyi5gPHbhkebPl1pLsCVrn4SpFe/wmbY8KjIT+8ZfOUn01X/RBcAcnz+/id2FVX+IekXSDNSSKNqVVX/cnBprq6SXIArWf1tba3rjaEPiqillOpXVRBTD3BsAmmZxC7Alu4Zlugog7j8ASjiMbXSZILs4bGxGyvJBbiS1a+Jbac8gM0w85jI4/cOvLwlmT6eRr0C5HjN/l+8XAB+mPQLpHYBEL7w7KL1DWvRbyvBBbhS1Z/Ntmxg5veLqE2Z+ZNAQaq96dX/dhcIxdwzLHbUpHaBSOZxuMyT+hsJ0EpwAa5Q9ROYTkn9ITOPqfznt3b/9qdp1X+8C6x+vf+3RdXvp3cBohG1qkRf2LvohopwAa5E9S9fueQaw+Z9p6R+BaB8GtT/dheIKPrqERuNGCCVCxQ0kiYKlrxhhz9TCS7Alah+EulRTTXPb1z9Yh/77uArPy8NI5/qiZVc4OJ9WweKpPc3sseU2gUiNYRbd8y/rnG2XaCSAsADIMvaFl9rjLkkWdollfqtKojSJ5AncwHD+rUjEg0zyJQ/IZV4DFaaOFhs4c26C1RSAAgAZkVq9QvUhsxcFP33+wcGtp4u9b+jRfDaz16NVP++kT2Cki3/ohMNS6REeuvO5g1zZ9MFuJLUn31363XG8MVp1U8gsqrKhnrP1In2Ji5gffM3RyQ6YghpbkrhAsTO47DlqKm7aTZdoFL6pBkAVrS1Pk3EF04VAIzxBZ0DHV/sOQDBV9hGNkasfXT14OBHnmxv90d37bJru7rGD9B/ssLtF5rGw6IUOUPI2+cWrbu3mYO/OCg2IpAXX05OLmu8mlBpnzC+Hx+ExIMhq/JGxprzVh7MH04KRM+2APAARG1tLdcTew+KyJSZ/wkCQANA68Bjvsj5DwwOvnIqJ1R6wMRU75eeQbBjwdUt4kUvkfIcmzhQGQEABaJ3UegdlOLGNQf+6SuPo8tbh/7obAsABkDZFa3PMPOaE9n/CQJA6kCsogMR4289JRMAGibR5QPwYBBKvB+g9HgYRgigHkbrIAcD9rZ9cmDHr5LWBJ/IDUqF9XzLlV9rovC2IYkigLwyA0B9GFjVA7YYnrdm6IFDM+0CXiWof8WKxd1s+AKxYlPO9mGrCmVkQ+Zvxev8T1j3P9kPiZLCT34rIwQhAyCAAUTw8LJVv2Sle2j3joeT5w1MGgSlAZ1XBPcepuJnDbgxioucpq8+oiIkepcJz3kLo39GwBfjGcQz5wJUAernFStanyXmjpMlf1M5QAgCx80I9RXWQykAxh0gfijEeM4QJscIS/sgCsHcBKY5xChauffqPTtuPdEDIkou8NzCq+5pNsEdByWKCOxN1wGSkWr1YRBB3kLI53buzh+cSRfgWVa/ZFe25tiYzlNo98fLtBwTFbzk2B5AHk2x4biNQIYAGlOxR8RGCzzv81uWrL47Xhl88vMquYASfeOwFA95KfoFCKAixDZxMF9H5eaZbhHMVgAQANvR0RGQ0qa07X5AQXH7H5LsnwYMAPOWjaJ64k2/bOm8pBuYNAhKw7rvff0/3rDAdxrZi5cKSPGdRyRSw/y5nS25c0o3q9ZyADAAHR09dDUzr06v/niyjz3NdRklI4kBEWD0tvjV3KSfLXXiBBx+c0iKBz1waheYS37z0aL+YewuM+MCsxIAXV1xeQlwLeJ7fFLXd6WbPU+/RZEZiZccWrezuX1uN/KT9taVLHvV3n89YEHfjl2AUriAUhGiRHpNHFgzc2fxrARAf3+cWZNqe7L2cmoBn8GrRJECBjR/pL5++YmS5pIL+Ca6b8gW3/RBptxqjUAUQYmAbNzR1Dcj08ZmtStYQf6pHuHMpsqqBiCOJAMA+SkKpOQCHa/95E2B3DcnpQsoFErwX+jAWZEEAqT7kgn7aYd+z2jpExEVoRECeivJAqb8ymMDOhn/20O2cMAnKmuMQKGIVynE/vNfzBc0vpFJazIASjkAKX4W1wTpIugM+6OG8am98uZg86CWlgSc+nx0C7rMBYOPHLTANxuo3BYBSQAGlJ6o+SSwvz++MCLFB0VkmCY+lqNCerAUauuICYp/XIf+aDoFUnIBGRv9zmEpHAhArJje4pUM4CgsFPQ9ANiPhTXdEaQAzODggb0q+iVj2KiiON0gKPW2n6krJNCojox/SKPdvu/fpwCtRf907iZSIMcXDfUfEujGRvaZgOhkVYFAC82c8Y6qvf+CA/mnZ/LxtLOZBAoAMzCw96vW2h95Hgdx2oUoeU+P3zR+aJMaJdVJ3k+76fhmVbVYT8ZT6EhB5OMXDTx3CGXUx5Q8U/CCNx79u7ds4bvzORMkwRopVMa/SwXxa3YBZ4KDduznljM3K3o4h/yMLSsz22MBpapcs22L/5qJPk9ERlUxWSsqA0KdxqNF49vEvt9kf8JgUGlM4PixgHgwaHwsIEOMOSA0ggHRHaJ64zV7fv3EyYaGp/q/FD1E6JMXF/7eJgPeWM9+OKaCKLnjwcAgJIOiKorQH0aKPz1/f/7ITCV/lRIAE89Bs9nFF5GhP4HqWgCtAPxkigUFSWHyJAFgjguA0mjgxEGheNLI2weDfCVkAM2ANAQNN4B3ZBQPvm6HvvfZvXtHUhb+xHqOCNBnWjZ01NvMjUq6ThRLFcQGvJ+VnrDgH3QcyPdP/PxsXPxKwEzImrm9vWV+oRB49QAakheHJ+wDQP0k+5O91jDh77d/rh6l49cH0fCndu06PKHwmE5DP9Px9fnO5g1zfRvyiqGHhyhuAkPRw0CfznThVyKMWZyj0APw4+jyTncPnKKHJ1spTJEzs/30UarQQKDTc+HL/tIzrsBScNGM9GU5HA6Hw+FwOBwOh8PhcDgcDofD4XA4HA6Hw+FwOBwOh8PhcDgcDofD4XA4apb/B36SpP07xdD5AAAAAElFTkSuQmCC"

# ===================== ระบบแปลภาษา TH / EN =====================
TEXT = {
    "th": {
        "page_title": "HR Smart Matrix Hub",
        "sidebar_logo_text": "HR Matrix Hub",
        "sidebar_step_label": "ขั้นตอนการทำงาน",
        "sidebar_step1_title": "เลือกเทมเพลตปลายทาง",
        "sidebar_step1_sub": "พร้อมใช้งานแล้ว",
        "sidebar_step2_title": "อัปโหลดไฟล์หลักฐาน",
        "sidebar_step2_sub_done": "อัปโหลดแล้ว: ",
        "sidebar_step2_sub_pending": "รองรับ Excel / รูปภาพ",
        "sidebar_step3_title": "วิเคราะห์ด้วย AI",
        "sidebar_step3_sub": "คัดแยกหมวดหมู่ + กรอกข้อมูลอัตโนมัติ",
        "sidebar_status_label": "สถานะระบบ",
        "sidebar_status_templates": "เทมเพลตที่พบ",
        "sidebar_status_templates_unit": "ไฟล์",
        "sidebar_status_model": "โมเดล AI",
        "sidebar_status_filetypes": "ไฟล์ที่รองรับ",
        "sidebar_templates_label": "เทมเพลตในระบบ",
        "sidebar_tips_label": "เคล็ดลับการใช้งาน",
        "sidebar_tip_text": "💡 ถ้าอัปโหลดรูปภาพ ให้ถ่ายให้เห็นตัวอักษรชัดเจนและไม่เอียง จะช่วยให้ AI อ่านข้อมูลได้แม่นยำขึ้น",
        "sidebar_pro_title": "⚡ Powered by Open-Source AI",
        "sidebar_pro_sub": "ระบบจับคู่และกรอกข้อมูลพนักงานอัตโนมัติ ลดเวลาทำงานซ้ำซ้อนของฝ่าย HR",
        "sidebar_lang_label": "🌐 ภาษา / Language",
        "main_header_title": "🗂️ HR Smart Matrix Hub",
        "main_header_sub": "อัปโหลดหลักฐาน ให้ AI คัดแยกหมวดหมู่ และจับคู่กรอกข้อมูลพนักงานลงเทมเพลตให้อัตโนมัติ — ไม่ต้องนั่งกรอกเอง",
        "step_pill_1": "เลือกเทมเพลตปลายทาง",
        "step_pill_2": "อัปโหลดไฟล์หลักฐาน",
        "step_pill_3": "ให้ AI วิเคราะห์และดาวน์โหลดผลลัพธ์",
        "step1_card_title": "ขั้นตอนที่ 1 · เลือกตารางแม่แบบปลายทาง",
        "step1_card_sub": "เลือกเทมเพลต Excel ที่ต้องการให้ระบบกรอกข้อมูลลงไป",
        "step1_select_label": "เลือกตารางแม่แบบปลายทาง:",
        "step1_columns_label": "📌 คอลัมน์ในเทมเพลตนี้: ",
        "step1_download_btn": "📥 ดาวน์โหลดเทมเพลตเปล่า",
        "step1_no_templates": "ไม่พบไฟล์แม่แบบในระบบหลังบ้าน",
        "step2_card_title": "ขั้นตอนที่ 2 · นำเข้าไฟล์หลักฐานดิบ",
        "step2_card_sub": "รองรับไฟล์ Excel (.xlsx, .xls) หรือรูปภาพเอกสาร (.png, .jpg, .jpeg)",
        "step2_uploader_label": "ลากและวางไฟล์หรือรูปภาพตรงนี้ หรือคลิกเพื่อเลือกไฟล์",
        "step2_image_caption": "รูปหลักฐานที่นำเข้าสู่ระบบ",
        "step2_file_ready": "✅ พร้อมวิเคราะห์ไฟล์: ",
        "step3_card_title": "ขั้นตอนที่ 3 · วิเคราะห์ด้วย AI และตรวจสอบผลลัพธ์",
        "step3_card_sub": "ระบบจะคัดแยกหมวดหมู่เอกสาร จับคู่ข้อมูล และกรอกลงเทมเพลตให้อัตโนมัติ",
        "step3_empty_title": "ยังไม่มีไฟล์ให้วิเคราะห์",
        "step3_empty_sub": "กรุณาอัปโหลดไฟล์หลักฐานในขั้นตอนที่ 2 ก่อน ระบบจะเปิดให้กดวิเคราะห์ที่นี่",
        "step3_analyze_btn": "🚀 สั่งเริ่มวิเคราะห์ข้อมูลด้วย AI",
        "step3_spinner": "🤖 AI กำลังอ่านข้อมูลและประมวลผล กรุณารอสักครู่...",
        "badge_personal": "🆔 คัดแยกสำเร็จ: หมวดหมู่ข้อมูลส่วนตัว / บัตรประชาชนพนักงาน",
        "badge_commission": "💰 คัดแยกสำเร็จ: หมวดหมู่รายได้สัมพันธ์ / ค่าคอมมิชชั่นพนักงาน",
        "badge_ot": "⏱️ คัดแยกสำเร็จ: หมวดหมู่ข้อมูลเวลาทำงาน / โอทีพนักงาน (OT)",
        "badge_generic": "📦 คัดแยกสำเร็จ: หมวดหมู่ทั่วไปในระบบ HR",
        "success_msg": "🎉 ประมวลผลและนำข้อมูลจัดระเบียบกรอกลงไฟล์ '{template_name}' สำเร็จแล้วค่ะ",
        "preview_heading": "#### 👀 ตารางตรวจสอบความถูกต้องก่อนบันทึก (Data Preview)",
        "download_result_btn": "📥 ดาวน์โหลดไฟล์ Excel ผลลัพธ์สมบูรณ์",
        "warning_no_match": "⚠️ ไม่สามารถบันทึกข้อมูลได้ เนื่องจากโครงสร้างข้อมูลไม่แมตช์กับหัวตารางในเทมเพลต",
        "error_processing": "❌ เกิดข้อผิดพลาดในการประมวลผลระบบ: ",
        "new_employee_label": "พนักงานรายใหม่",
        "col_row_index": "แถวตารางที่",
        "col_full_name": "ชื่อ-นามสกุล",
    },
    "en": {
        "page_title": "HR Smart Matrix Hub",
        "sidebar_logo_text": "HR Matrix Hub",
        "sidebar_step_label": "Workflow Steps",
        "sidebar_step1_title": "Select destination template",
        "sidebar_step1_sub": "Ready to use",
        "sidebar_step2_title": "Upload source file",
        "sidebar_step2_sub_done": "Uploaded: ",
        "sidebar_step2_sub_pending": "Supports Excel / Images",
        "sidebar_step3_title": "Analyze with AI",
        "sidebar_step3_sub": "Classify category + auto-fill data",
        "sidebar_status_label": "System Status",
        "sidebar_status_templates": "Templates found",
        "sidebar_status_templates_unit": "file(s)",
        "sidebar_status_model": "AI Model",
        "sidebar_status_filetypes": "Supported files",
        "sidebar_templates_label": "Templates in system",
        "sidebar_tips_label": "Usage Tips",
        "sidebar_tip_text": "💡 If uploading a photo, make sure the text is clear and not tilted — this helps the AI read the data more accurately.",
        "sidebar_pro_title": "⚡ Powered by Open-Source AI",
        "sidebar_pro_sub": "Automatically matches and fills employee data, reducing repetitive HR workload.",
        "sidebar_lang_label": "🌐 ภาษา / Language",
        "main_header_title": "🗂️ HR Smart Matrix Hub",
        "main_header_sub": "Upload your source document, let AI classify it, and automatically match and fill employee data into the template — no manual entry needed.",
        "step_pill_1": "Select destination template",
        "step_pill_2": "Upload source file",
        "step_pill_3": "Analyze with AI & download result",
        "step1_card_title": "Step 1 · Select destination template",
        "step1_card_sub": "Choose the Excel template you want the system to fill data into",
        "step1_select_label": "Select destination template:",
        "step1_columns_label": "📌 Columns in this template: ",
        "step1_download_btn": "📥 Download blank template",
        "step1_no_templates": "No template files found on the backend",
        "step2_card_title": "Step 2 · Import source document",
        "step2_card_sub": "Supports Excel files (.xlsx, .xls) or document images (.png, .jpg, .jpeg)",
        "step2_uploader_label": "Drag and drop a file or image here, or click to browse",
        "step2_image_caption": "Source image imported into the system",
        "step2_file_ready": "✅ Ready to analyze file: ",
        "step3_card_title": "Step 3 · AI analysis & result review",
        "step3_card_sub": "The system will classify the document, match the data, and fill it into the template automatically",
        "step3_empty_title": "No file to analyze yet",
        "step3_empty_sub": "Please upload a source file in Step 2 first — the analyze button will appear here.",
        "step3_analyze_btn": "🚀 Start AI Analysis",
        "step3_spinner": "🤖 AI is reading and processing the data, please wait...",
        "badge_personal": "🆔 Classified: Personal Data / Employee ID Card",
        "badge_commission": "💰 Classified: Commission / Employee Income",
        "badge_ot": "⏱️ Classified: Working Hours / Employee Overtime (OT)",
        "badge_generic": "📦 Classified: General HR Category",
        "success_msg": "🎉 Data processed and filled into '{template_name}' successfully!",
        "preview_heading": "#### 👀 Data Preview (Verify before saving)",
        "download_result_btn": "📥 Download Completed Excel File",
        "warning_no_match": "⚠️ Could not save data because the structure does not match the template headers.",
        "error_processing": "❌ An error occurred while processing: ",
        "new_employee_label": "New Employee",
        "col_row_index": "Row No.",
        "col_full_name": "Full Name",
    },
}

if "lang" not in st.session_state:
    st.session_state["lang"] = "th"

# 1. ตั้งค่าหน้าเว็บระดับพรีเมียม
st.set_page_config(
    page_title="HR Smart Matrix Hub",
    page_icon="🗂️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===================== CSS THEME: Dark / Indigo =====================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .stApp, [data-testid="stAppViewContainer"] {
        background-color: #0D0D14;
    }
    [data-testid="stHeader"] { background-color: rgba(0,0,0,0); }

    [data-testid="stSidebar"] {
        background-color: #11111A;
        border-right: 1px solid #23232F;
    }
    [data-testid="stSidebar"] * { color: #CBD0DC; }
    
    .sb-logo { display: flex; align-items: center; gap: 10px; padding: 6px 0 18px 0; margin-bottom: 6px; border-bottom: 1px solid #23232F; }
    .sb-logo-icon { width: 34px; height: 34px; border-radius: 9px; background: #FFFFFF; display: flex; align-items: center; justify-content: center; font-size: 1.05rem; flex-shrink: 0; padding: 5px; box-sizing: border-box; }
    .sb-logo-icon img { width: 100%; height: 100%; object-fit: contain; }
    .sb-logo-text { font-size: 1.15rem; font-weight: 800; color: #F4F4F8; letter-spacing: -0.3px; }
    .sb-section-label { font-size: 0.72rem; font-weight: 700; letter-spacing: 0.06em; color: #5C5F70; text-transform: uppercase; margin: 22px 0 10px 0; }
    .sb-step { display: flex; align-items: flex-start; gap: 10px; padding: 10px 12px; border-radius: 10px; margin-bottom: 6px; background: #15151F; border: 1px solid #23232F; }
    .sb-step.active { background: rgba(109, 93, 251, 0.12); border: 1px solid rgba(109, 93, 251, 0.4); }
    .sb-step.done { background: rgba(34, 197, 94, 0.07); border: 1px solid rgba(34, 197, 94, 0.25); }
    .sb-step-num { display: flex; align-items: center; justify-content: center; width: 20px; height: 20px; border-radius: 50%; background: #2A2A38; color: #9799AB; font-size: 0.7rem; font-weight: 700; flex-shrink: 0; margin-top: 1px; }
    .sb-step.active .sb-step-num { background: #6D5DFB; color: white; }
    .sb-step.done .sb-step-num { background: #22C55E; color: white; }
    .sb-step-text { font-size: 0.85rem; font-weight: 600; color: #E4E5EC; line-height: 1.3; }
    .sb-step-sub { font-size: 0.74rem; color: #777A8C; margin-top: 2px; line-height: 1.3; }
    .sb-info-card { background: #15151F; border: 1px solid #23232F; border-radius: 10px; padding: 14px; margin-bottom: 8px; }
    .sb-info-row { display: flex; justify-content: space-between; font-size: 0.82rem; padding: 4px 0; color: #B6B8C6; }
    .sb-info-row b { color: #F4F4F8; }
    .sb-tip { font-size: 0.78rem; color: #9799AB; line-height: 1.5; background: #15151F; border: 1px solid #23232F; border-radius: 10px; padding: 12px 14px; }
    .sb-pro-card { background: linear-gradient(160deg, #2A1F6B 0%, #6D5DFB 100%); border-radius: 12px; padding: 18px; margin-top: 22px; color: white; }
    .sb-pro-card b { color: white; font-size: 0.95rem; }
    .sb-pro-card p { font-size: 0.78rem; opacity: 0.85; margin: 6px 0 0 0; }

    .main-header { background: linear-gradient(135deg, #15151F 0%, #1B1730 100%); padding: 32px 36px; border-radius: 16px; color: #F4F4F8; margin-bottom: 26px; border: 1px solid #23232F; position: relative; overflow: hidden; }
    .main-header::after { content: ""; position: absolute; top: -50%; right: -8%; width: 280px; height: 280px; background: radial-gradient(circle, rgba(109,93,251,0.18) 0%, rgba(109,93,251,0) 70%); border-radius: 50%; }
    .main-header h1 { margin: 0; font-size: 1.95rem; font-weight: 800; letter-spacing: -0.5px; color: #F4F4F8; }
    .main-header p { margin: 8px 0 0 0; opacity: 0.7; font-size: 0.98rem; font-weight: 300; color: #C7C9D6; }

    .step-track { display: flex; align-items: center; gap: 8px; margin-bottom: 26px; flex-wrap: wrap; }
    .step-pill { display: flex; align-items: center; gap: 8px; background: #15151F; border: 1px solid #23232F; border-radius: 999px; padding: 8px 18px 8px 10px; font-size: 0.86rem; font-weight: 600; color: #777A8C; }
    .step-pill.active { background: rgba(109, 93, 251, 0.12); border-color: rgba(109, 93, 251, 0.45); color: #C9C2FF; }
    .step-pill.done { background: rgba(34, 197, 94, 0.08); border-color: rgba(34, 197, 94, 0.3); color: #86EFAC; }
    .step-num { display: flex; align-items: center; justify-content: center; width: 22px; height: 22px; border-radius: 50%; background: #2A2A38; color: #C7C9D6; font-size: 0.75rem; font-weight: 700; }
    .step-pill.active .step-num { background: #6D5DFB; color: white; }
    .step-pill.done .step-num { background: #22C55E; color: white; }
    .step-connector { width: 24px; height: 2px; background: #23232F; }

    .card-box { background-color: #14141E; padding: 26px 28px; border-radius: 14px; border: 1px solid #23232F; box-shadow: 0 4px 20px rgba(0, 0, 0, 0.25); margin-bottom: 22px; }
    .card-box-header { display: flex; align-items: center; gap: 12px; margin-bottom: 18px; }
    .card-box-icon { display: flex; align-items: center; justify-content: center; width: 38px; height: 38px; border-radius: 10px; background: rgba(109, 93, 251, 0.15); font-size: 1.1rem; flex-shrink: 0; }
    .card-box-title { font-size: 1.05rem; font-weight: 700; color: #F4F4F8; margin: 0; }
    .card-box-subtitle { font-size: 0.85rem; color: #777A8C; margin: 2px 0 0 0; }

    .badge-personal, .badge-commission, .badge-ot, .badge-generic { padding: 10px 20px; border-radius: 10px; font-weight: 600; font-size: 0.95rem; display: inline-flex; align-items: center; gap: 8px; }
    .badge-personal { background-color: rgba(248, 113, 113, 0.1); color: #FCA5A5; border: 1px solid rgba(248, 113, 113, 0.3); }
    .badge-commission { background-color: rgba(74, 222, 128, 0.1); color: #86EFAC; border: 1px solid rgba(74, 222, 128, 0.3); }
    .badge-ot { background-color: rgba(251, 191, 36, 0.1); color: #FCD34D; border: 1px solid rgba(251, 191, 36, 0.3); }
    .badge-generic { background-color: rgba(109, 93, 251, 0.12); color: #C9C2FF; border: 1px solid rgba(109, 93, 251, 0.35); }

    div.stButton > button:first-child { background: linear-gradient(135deg, #6D5DFB 0%, #8B7CFF 100%) !important; color: white !important; border-radius: 10px !important; font-size: 16px !important; font-weight: 600 !important; padding: 14px 30px !important; border: none !important; width: 100% !important; box-shadow: 0 4px 16px rgba(109, 93, 251, 0.3) !important; transition: all 0.2s ease !important; }
    div.stButton > button:first-child:hover { background: linear-gradient(135deg, #5B4BF0 0%, #7A6BFF 100%) !important; transform: translateY(-1px) !important; box-shadow: 0 6px 20px rgba(109, 93, 251, 0.4) !important; }
    div.stDownloadButton > button:first-child { background: #1A1A26 !important; color: #C9C2FF !important; border: 1.5px solid rgba(109, 93, 251, 0.4) !important; border-radius: 10px !important; font-weight: 600 !important; padding: 12px 24px !important; width: 100% !important; transition: all 0.2s ease !important; }
    div.stDownloadButton > button:first-child:hover { background: rgba(109, 93, 251, 0.12) !important; border-color: #6D5DFB !important; }

    [data-testid="stFileUploader"] section { background-color: #15151F !important; border: 1.5px dashed #2A2A38 !important; border-radius: 12px !important; }
    [data-testid="stFileUploader"] small { color: #777A8C !important; }
    div[data-baseweb="select"] > div { background-color: #15151F !important; border-color: #2A2A38 !important; color: #E4E5EC !important; }
    .stTextInput input, .stTextArea textarea { background-color: #15151F !important; color: #E4E5EC !important; border-color: #2A2A38 !important; }
    h1, h2, h3, h4, h5, p, span, label, .stMarkdown { color: #E4E5EC; }
    .helper-text { font-size: 0.85rem; color: #777A8C; margin-top: 6px; }
    .divider-line { border-top: 1px dashed #23232F; margin: 22px 0; }
    .empty-state { text-align: center; padding: 50px 20px; color: #5C5F70; }
    .empty-state-icon { font-size: 2.8rem; margin-bottom: 12px; opacity: 0.5; }
    div[data-testid="stNotification"] { background-color: #15151F !important; border: 1px solid #23232F !important; border-radius: 10px !important; }
    </style>
""", unsafe_allow_html=True)

TEMPLATES_FOLDER = "hr_templates"

def learn_templates():
    structures = {}
    if not os.path.exists(TEMPLATES_FOLDER):
        return structures
    for filename in os.listdir(TEMPLATES_FOLDER):
        if filename.endswith((".xlsx", ".xls")) and not filename.startswith((".", "~$")):
            file_path = os.path.join(TEMPLATES_FOLDER, filename)
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() for cell in ws[1] if cell.value is not None]

                existing_rows = []
                for row_idx in range(2, ws.max_row + 1):
                    row_data = {}
                    has_data = False
                    for col_idx, header in enumerate(headers, 1):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        val_str = str(val).strip() if val is not None else ""
                        row_data[header] = val_str
                        if val_str: has_data = True
                    if has_data:
                        row_data["__row_index__"] = row_idx
                        existing_rows.append(row_data)

                if headers:
                    structures[filename] = {"headers": headers, "existing_data": existing_rows}
            except:
                pass
    return structures

templates_available = learn_templates()

_uploaded_marker = st.session_state.get("_last_uploaded_name")
step2_state = "done" if _uploaded_marker else "active"
step3_state = "active" if _uploaded_marker else ""

# ===================== SIDEBAR =====================
with st.sidebar:
    lang_options = {"th": "🇹🇭 ไทย", "en": "🇬🇧 English"}
    selected_lang_label = st.radio(
        TEXT[st.session_state["lang"]]["sidebar_lang_label"],
        options=list(lang_options.values()),
        index=list(lang_options.keys()).index(st.session_state["lang"]),
        horizontal=True,
        label_visibility="visible",
        key="lang_radio"
    )
    st.session_state["lang"] = [k for k, v in lang_options.items() if v == selected_lang_label][0]

    lang = st.session_state["lang"]
    T = TEXT[lang]

    st.markdown(f"""
        <div class="sb-logo">
            <div class="sb-logo-icon"><img src="data:image/png;base64,{LOGO_BASE64}" alt="AREE Workforce Tech logo"/></div>
            <div class="sb-logo-text">{T['sidebar_logo_text']}</div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown(f'<div class="sb-section-label">{T["sidebar_step_label"]}</div>', unsafe_allow_html=True)
    step2_sub_text = (T["sidebar_step2_sub_done"] + _uploaded_marker) if _uploaded_marker else T["sidebar_step2_sub_pending"]
    st.markdown(f"""
        <div class="sb-step done">
            <div class="sb-step-num">1</div>
            <div>
                <div class="sb-step-text">{T['sidebar_step1_title']}</div>
                <div class="sb-step-sub">{T['sidebar_step1_sub']}</div>
            </div>
        </div>
        <div class="sb-step {step2_state}">
            <div class="sb-step-num">2</div>
            <div>
                <div class="sb-step-text">{T['sidebar_step2_title']}</div>
                <div class="sb-step-sub">{step2_sub_text}</div>
            </div>
        </div>
        <div class="sb-step {step3_state}">
            <div class="sb-step-num">3</div>
            <div>
                <div class="sb-step-text">{T['sidebar_step3_title']}</div>
                <div class="sb-step-sub">{T['sidebar_step3_sub']}</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown(f'<div class="sb-section-label">{T["sidebar_status_label"]}</div>', unsafe_allow_html=True)
    st.markdown(f"""
        <div class="sb-info-card">
            <div class="sb-info-row"><span>{T['sidebar_status_templates']}</span><b>{len(templates_available)} {T['sidebar_status_templates_unit']}</b></div>
            <div class="sb-info-row"><span>{T['sidebar_status_model']}</span><b>Ollama (Local LLM)</b></div>
            <div class="sb-info-row"><span>{T['sidebar_status_filetypes']}</span><b>XLSX · PNG · JPG</b></div>
        </div>
    """, unsafe_allow_html=True)

    if templates_available:
        st.markdown(f'<div class="sb-section-label">{T["sidebar_templates_label"]}</div>', unsafe_allow_html=True)
        for tname in templates_available.keys():
            st.markdown(f"<div class='sb-tip' style='margin-bottom:6px;'>📄 {tname}</div>", unsafe_allow_html=True)

    st.markdown(f'<div class="sb-section-label">{T["sidebar_tips_label"]}</div>', unsafe_allow_html=True)
    st.markdown(f"""
        <div class="sb-tip">
            {T['sidebar_tip_text']}
        </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
        <div class="sb-pro-card">
            <b>{T['sidebar_pro_title']}</b>
            <p>{T['sidebar_pro_sub']}</p>
        </div>
    """, unsafe_allow_html=True)

# ===================== MAIN UI =====================
st.markdown(f"""
    <div class="main-header">
        <h1>{T['main_header_title']}</h1>
        <p>{T['main_header_sub']}</p>
    </div>
""", unsafe_allow_html=True)

st.markdown(f"""
    <div class="step-track">
        <div class="step-pill done"><span class="step-num">1</span>{T['step_pill_1']}</div>
        <div class="step-connector"></div>
        <div class="step-pill {step2_state}"><span class="step-num">2</span>{T['step_pill_2']}</div>
        <div class="step-connector"></div>
        <div class="step-pill {step3_state}"><span class="step-num">3</span>{T['step_pill_3']}</div>
    </div>
""", unsafe_allow_html=True)


# --- STEP 1 ---
st.markdown('<div class="card-box">', unsafe_allow_html=True)
st.markdown(f"""
    <div class="card-box-header">
        <div class="card-box-icon">📁</div>
        <div>
            <p class="card-box-title">{T['step1_card_title']}</p>
            <p class="card-box-subtitle">{T['step1_card_sub']}</p>
        </div>
    </div>
""", unsafe_allow_html=True)

if templates_available:
    col_select, col_download = st.columns([2, 1], gap="medium")
    with col_select:
        selected_t_download = st.selectbox(T["step1_select_label"], list(templates_available.keys()), label_visibility="collapsed")
        st.markdown(f"<p class='helper-text'>{T['step1_columns_label']}{', '.join(templates_available[selected_t_download]['headers'])}</p>", unsafe_allow_html=True)

    t_download_path = os.path.join(TEMPLATES_FOLDER, selected_t_download)
    with open(t_download_path, "rb") as f:
        with col_download:
            st.download_button(
                label=T["step1_download_btn"],
                data=f,
                file_name=selected_t_download,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
else:
    st.error(T["step1_no_templates"])
st.markdown('</div>', unsafe_allow_html=True)


# --- STEP 2 ---
st.markdown('<div class="card-box">', unsafe_allow_html=True)
st.markdown(f"""
    <div class="card-box-header">
        <div class="card-box-icon">📤</div>
        <div>
            <p class="card-box-title">{T['step2_card_title']}</p>
            <p class="card-box-subtitle">{T['step2_card_sub']}</p>
        </div>
    </div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    T["step2_uploader_label"],
    type=["xlsx", "xls", "png", "jpg", "jpeg"],
    label_visibility="visible"
)

if uploaded_file is not None:
    st.session_state["_last_uploaded_name"] = uploaded_file.name
    file_ext = uploaded_file.name.split(".")[-1].lower()
    is_image = file_ext in ["png", "jpg", "jpeg"]
    if is_image:
        col_img, col_spacer = st.columns([1, 1], gap="medium")
        with col_img:
            image_obj = Image.open(uploaded_file)
            st.image(image_obj, caption=T["step2_image_caption"], use_container_width=True)
    else:
        st.markdown(f"<p class='helper-text'>{T['step2_file_ready']}<b>{uploaded_file.name}</b></p>", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)


# --- STEP 3 ---
st.markdown('<div class="card-box">', unsafe_allow_html=True)
st.markdown(f"""
    <div class="card-box-header">
        <div class="card-box-icon">🤖</div>
        <div>
            <p class="card-box-title">{T['step3_card_title']}</p>
            <p class="card-box-subtitle">{T['step3_card_sub']}</p>
        </div>
    </div>
""", unsafe_allow_html=True)

if uploaded_file is None:
    st.markdown(f"""
        <div class="empty-state">
            <div class="empty-state-icon">🗒️</div>
            <p style="margin:0; font-weight:600; color:#9799AB;">{T['step3_empty_title']}</p>
            <p style="margin:4px 0 0 0; font-size:0.9rem;">{T['step3_empty_sub']}</p>
        </div>
    """, unsafe_allow_html=True)
else:
    if st.button(T["step3_analyze_btn"], use_container_width=True):
        with st.spinner(T["step3_spinner"]):
            try:
                # 1. การดึงข้อมูลต้นฉบับ (OCR จากรูปภาพ หรือ แกะข้อมูลตารางจาก Excel)
                if is_image:
                    uploaded_file.seek(0)
                    extracted_text = pytesseract.image_to_string(Image.open(uploaded_file), lang='tha+eng')
                    document_data = f"[ข้อมูลจากรูปภาพที่สแกนได้]\n{extracted_text}"
                else:
                    wb_in = load_workbook(uploaded_file, data_only=True)
                    input_text = ""
                    for sheet_name in wb_in.sheetnames:
                        ws_in = wb_in[sheet_name]
                        input_text += f"\n[หน้าตาราง: {sheet_name}]\n"
                        for row in ws_in.iter_rows(values_only=True):
                            row_clean = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
                            if row_clean:
                                input_text += " | ".join(row_clean) + "\n"
                    document_data = f"[ข้อมูลเอกสารดิบ]\n{input_text}"

                # 2. การสร้าง Prompt เตรียมส่งให้ AI
                blueprint = json.dumps(templates_available, ensure_ascii=False, indent=2)
                prompt = f"""
                คุณคือผู้เชี่ยวชาญด้านระบบ HR Automation หน้าที่ของคุณคือวิเคราะห์ข้อมูลพนักงานจากเอกสารดิบ
                แล้วนำข้อมูลไปจับคู่เติมลงในตารางพนักงานของแต่ละเทมเพลตในจุดที่ข้อมูลยังว่างอยู่ให้ถูกต้องและสมบูรณ์

                นี่คือโครงสร้างตารางและรายชื่อพนักงานปัจจุบันที่มีอยู่ในระบบแยกตามไฟล์เทมเพลตต่างๆ:
                {blueprint}

                คำสั่งในการวิเคราะห์และจัดหมวดหมู่:
                1. ระบุประเภทหมวดหมู่ (Category) ของหลักฐานนี้ โดยเลือกค่าต่อไปนี้: "PERSONAL_DATA", "COMMISSION", "OVERTIME", "GENERIC_HR"
                2. คัดเลือกเทมเพลตปลายทางที่เหมาะสมที่สุดกรอกลงช่อง "selected_template"
                3. ตรวจสอบรหัสหรือชื่อเพื่อทำการจับคู่ หากในตารางปลายทางช่องรายชื่อยังไม่มีค่า ให้ดึงข้อมูล "ชื่อ-นามสกุล" ที่สแกนได้จากหลักฐานมาเติมใส่ด้วย

                ให้ตอบกลับเป็นรูปแบบ JSON โครงสร้างแบบนี้เท่านั้น ห้ามมีคำอธิบายอื่นเด็ดขาด:
                {{
                  "category": "หมวดหมู่เอกสาร",
                  "selected_template": "ชื่อไฟล์เทมเพลตปลายทาง.xlsx",
                  "updates": [
                      {{
                        "row_index": 2,
                        "data_to_fill": {{
                           "ชื่อคอลัมน์": "ค่าที่ดึงได้"
                        }}
                      }}
                  ]
                }}
                """

                # ผสมคำสั่งเข้ากับข้อมูลเอกสารที่แกะมาได้
                full_prompt = prompt + "\n\n" + document_data

                # 3. ส่งไปให้ Local Ollama Model แทน Gemini
                url = "http://localhost:11434/api/generate"
                data = {
                    "model": "typhoon-m1", # เปลี่ยนเป็น llama3.1 หรือโมเดลที่คุณใช้งานอยู่ได้
                    "prompt": full_prompt,
                    "stream": False,
                    "format": "json"
                }
                
                response = requests.post(url, json=data)
                
                if response.status_code == 200:
                    result_json_str = response.json().get('response', '{}')
                    ai_result = json.loads(result_json_str)
                else:
                    st.error("❌ ไม่สามารถเชื่อมต่อกับ AI (Ollama) ได้ กรุณาตรวจสอบว่าเปิดโปรแกรม Ollama และโหลด Model ไว้หรือไม่")
                    st.stop()

                # 4. ประมวลผลผลลัพธ์ที่ได้จาก AI (เหมือนเดิม 100%)
                category = ai_result.get("category", "GENERIC_HR")
                template_name = ai_result.get("selected_template")
                updates = ai_result.get("updates", [])

                st.markdown('<div class="divider-line"></div>', unsafe_allow_html=True)
                if category == "PERSONAL_DATA":
                    st.markdown(f'<div class="badge-personal">{T["badge_personal"]}</div>', unsafe_allow_html=True)
                elif category == "COMMISSION":
                    st.markdown(f'<div class="badge-commission">{T["badge_commission"]}</div>', unsafe_allow_html=True)
                elif category == "OVERTIME":
                    st.markdown(f'<div class="badge-ot">{T["badge_ot"]}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="badge-generic">{T["badge_generic"]}</div>', unsafe_allow_html=True)
                st.write("")

                template_path = os.path.join(TEMPLATES_FOLDER, template_name)
                wb_target = load_workbook(template_path)
                ws_target = wb_target.active

                header_map = {str(ws_target.cell(row=1, column=c).value).strip(): c for c in range(1, ws_target.max_column + 1)}
                row_preview_list = []

                for item in updates:
                    r_idx = item.get("row_index")
                    data_to_fill = item.get("data_to_fill", {})

                    emp_name_col = header_map.get("รายชื่อ") or header_map.get("ชื่อพนักงาน") or header_map.get("ชื่อ") or 1
                    existing_name = ws_target.cell(row=r_idx, column=emp_name_col).value

                    display_name = existing_name if existing_name else data_to_fill.get("รายชื่อ") or data_to_fill.get("ชื่อพนักงาน") or data_to_fill.get("ชื่อ") or T["new_employee_label"]

                    preview_row = {T["col_row_index"]: r_idx, T["col_full_name"]: display_name}

                    for h_name, new_val in data_to_fill.items():
                        if h_name in header_map:
                            ws_target.cell(row=r_idx, column=header_map[h_name]).value = new_val
                            preview_row[h_name] = new_val

                    row_preview_list.append(preview_row)

                if row_preview_list:
                    st.balloons()
                    st.success(T["success_msg"].replace("{template_name}", template_name))
                    st.markdown(T["preview_heading"])
                    df_wide = pd.DataFrame(row_preview_list)

                    cols = list(df_wide.columns)
                    if T["col_row_index"] in cols: cols.insert(0, cols.pop(cols.index(T["col_row_index"])))
                    if T["col_full_name"] in cols: cols.insert(1, cols.pop(cols.index(T["col_full_name"])))
                    df_wide = df_wide[cols]

                    st.dataframe(
                        df_wide.style.set_properties(**{
                            'background-color': '#14141E',
                            'color': '#E4E5EC',
                            'border-color': '#23232F'
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                    out_stream = io.BytesIO()
                    wb_target.save(out_stream)
                    out_stream.seek(0)

                    st.write("")
                    st.download_button(
                        label=T["download_result_btn"],
                        data=out_stream,
                        file_name=f"HR_SUCCESS_{template_name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.warning(T["warning_no_match"])

            except Exception as e:
                st.error(f"{T['error_processing']} {e}")
st.markdown('</div>', unsafe_allow_html=True)
